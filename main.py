from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl
import re


def get_html(url):
    driver = webdriver.Chrome('chromedriver')
    message = set()
    try:
        driver.get(url=url)
        input()
        # soup = BeautifulSoup(driver.page_source, 'html.parser')
        for i in range(1, 10):

            find_more_element = driver.find_element(By.XPATH,
                                                    '/html/body/div[2]/div[1]/div[2]/div/div[2]/div[3]/div/div/div/div[2]')
            actions = ActionChains(driver)
            actions.move_to_element(find_more_element).perform()
            time.sleep(0.3)

            soup = BeautifulSoup(driver.page_source, 'html.parser')
            try:
                all_message = soup.find_all('div', {"class": 'bubble'})
                for ms in all_message:
                    try:
                        message.add(ms)
                    except:
                        print('Error')
                print(len(message))
            except:
                print('Error')
        add_to_exel(message)
        #         if len(message) > 500:
        #             add_to_exel(message)
        #             message = set()
        #
        #             print('Продолжить - ? Y/N')
        #             process = str(input().lower())
        #             if process == 'n':
        #                 break
        #
        #     except:
        #         print('Error')
        #
        # # add_to_exel(message)
        #
        # print('*'*4)
        input()

    except Exception as ex:
        print(ex)
    finally:
        driver.close()
        driver.quit()


def add_to_exel(message_obj):
    excel_obj = openpyxl.load_workbook('Sheat.xlsx')
    ws = excel_obj.active
    for ms in message_obj:
        try:
            user_id = ms.find('div', {"class": "name"}).get('data-peer-id')
        except:
            user_id = 'Not_found'
        try:
            message_data = ms.find('span', {"class": 'time'}).get('title')
            message_data_text = str(message_data)
            if message_data.find('Edited') != -1:
                message_data_edited_index = message_data.find('Edited')
                message_data_text = message_data[:message_data_edited_index]
                if message_data.find('Original') != -1:
                    message_data_original = message_data[message_data.find('Original'):]
                    message_data_edited = message_data[message_data_edited_index:message_data.find('Original')]
                else:
                    message_data_edited = message_data[message_data_edited_index:]
                    message_data_original = 'Not_found'
            elif message_data.find('Original') != -1:
                message_data_original = message_data[message_data.find('Original'):]
                message_data_text = message_data[:message_data.find('Original')]
                message_data_edited = 'Not_found'
            else:
                message_data_edited = 'Not_found'
                message_data_original = 'Not_found'
        except:
            message_data = 'Not_found'
            message_data_text = 'Not_found'
            message_data_edited = 'Not_found'
            message_data_original = 'Not_found'
        try:
            name = ms.find('div', {"class": "name"}).get_text()
        except:
            name = 'Not_found'
        try:
            article = ms.find('div', {"class": 'message'}).get_text()
        except:
            article = 'Not_found'
        try:
            reply_name_id = ms.find('div', {"class": 'reply-title'}).find('span', {"class": "peer-title"}).get('data-peer-id')
        except:
            reply_name_id = 'Not_found'
        try:
            reply_name = ms.find('div', {"class": 'reply-title'}).find('span', {"class": "peer-title"}).get_text()
        except:
            reply_name = 'Not_found'
        try:
            reply_article = ms.find('div', {"class": "reply-subtitle"}).get_text()
        except:
            reply_article = 'Not_found'

        ws[f'A{str(ws.max_row+1)}'] = user_id
        ws[f'B{str(ws.max_row)}'] = name
        ws[f'C{str(ws.max_row)}'] = article
        ws[f'D{str(ws.max_row)}'] = reply_name_id
        ws[f'E{str(ws.max_row)}'] = reply_name
        ws[f'F{str(ws.max_row)}'] = reply_article
        ws[f'G{str(ws.max_row)}'] = message_data_text
        ws[f'H{str(ws.max_row)}'] = re.sub(r'\w+\: ', '', message_data_edited)
        ws[f'I{str(ws.max_row)}'] = re.sub(r'\w+\: ', '', message_data_original)
        print()
        print(message_data)
        # ws[f'C{str(ws.max_row)}'] = message_author
        # ws[f'D{str(ws.max_row)}'] = message_time

    excel_obj.save('Sheat.xlsx')


if __name__ == '__main__':
    get_html('https://xn--80affa3aj0al.xn--80asehdb/')

    # parsing_telegram()
